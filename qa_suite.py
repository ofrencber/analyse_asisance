from __future__ import annotations

import io
import os
import sys
import traceback
from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest


ROOT = Path(__file__).resolve().parent
os.chdir(ROOT)
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import mcdm_engine as me


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def make_dataset(seed: int, *, mixed: bool = False, n_alt: int = 8, n_crit: int = 5) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    data: dict[str, np.ndarray] = {}
    for idx in range(n_crit):
        loc = 60.0 + idx * 8.0
        scale = 10.0 + idx * 2.5
        col = rng.normal(loc=loc, scale=scale, size=n_alt)
        if mixed:
            col = col - (55.0 + idx * 4.0)
        data[f"C{idx + 1}"] = col
    index = [f"A{i + 1:02d}" for i in range(n_alt)]
    return pd.DataFrame(data, index=index)


def criteria_types_for(columns: list[str]) -> dict[str, str]:
    return {col: ("max" if idx % 2 == 0 else "min") for idx, col in enumerate(columns)}


def assert_no_streamlit_exceptions(at: AppTest, context: str) -> None:
    errors = list(at.exception)
    if errors:
        raise AssertionError(f"{context}: Streamlit exception detected: {errors[0]}")


def make_app_test(timeout: int = 120) -> AppTest:
    at = AppTest.from_file(str(ROOT / "mcdm_app.py"), default_timeout=timeout)
    # Bare-mode testlerde auth duvarını kapatıp asıl analiz akışını doğruluyoruz.
    at.secrets["mcdm_auth"] = {"require_login": False}
    return at


def session_state_value(at: AppTest, key: str, default=None):
    return at.session_state.filtered_state.get(key, default)


def click_button_by_key(at: AppTest, key: str) -> None:
    at.button(key=key).click().run()
    assert_no_streamlit_exceptions(at, f"button:{key}")


def click_button_by_label(at: AppTest, *needles: str) -> None:
    wanted = [needle.lower() for needle in needles]
    for button in at.button:
        label = str(getattr(button, "label", "") or "").lower()
        if any(needle in label for needle in wanted):
            button.click().run()
            assert_no_streamlit_exceptions(at, f"button:{label}")
            return
    raise AssertionError(f"Button not found for labels: {needles}")


def set_checkbox(at: AppTest, key: str, value: bool = True) -> None:
    at.checkbox(key=key).set_value(value).run()
    assert_no_streamlit_exceptions(at, f"checkbox:{key}")


def set_radio_by_label(at: AppTest, label_needles: list[str], value: str) -> None:
    wanted = [needle.lower() for needle in label_needles]
    for radio in at.radio:
        label = str(getattr(radio, "label", "") or "").lower()
        if any(needle in label for needle in wanted):
            radio.set_value(value).run()
            assert_no_streamlit_exceptions(at, f"radio:{label}")
            return
    raise AssertionError(f"Radio not found for labels: {label_needles}")


def get_download_workbook(at: AppTest) -> tuple[bytes, list[str]]:
    cache = dict(at.session_state["download_blob_cache"])
    excel_keys = [key for key in cache if str(key).startswith("excel::")]
    require(excel_keys, "No Excel export blob found in session state.")
    blob = cache[excel_keys[0]]
    require(isinstance(blob, (bytes, bytearray)) and len(blob) > 0, "Excel export blob is empty.")
    wb = load_workbook(io.BytesIO(blob))
    return bytes(blob), wb.sheetnames


def test_weight_methods() -> None:
    for mixed in (False, True):
        df = make_dataset(101 if not mixed else 202, mixed=mixed)
        criteria = list(df.columns)
        types = criteria_types_for(criteria)
        for method in me.OBJECTIVE_WEIGHT_METHODS:
            weights, details = me.compute_objective_weights(df, criteria, types, method, fuzzy_spread=0.12)
            arr = np.asarray([weights[c] for c in criteria], dtype=float)
            require(np.isfinite(arr).all(), f"{method}: non-finite weights.")
            require(np.all(arr >= -1e-10), f"{method}: negative weights produced.")
            require(np.isclose(arr.sum(), 1.0, atol=1e-8), f"{method}: weights do not sum to 1.")
            require(isinstance(details.get("weight_table"), pd.DataFrame), f"{method}: weight table missing.")


def test_ranking_methods() -> None:
    positive_df = make_dataset(303, mixed=False)
    mixed_df = make_dataset(404, mixed=True)
    criteria = list(positive_df.columns)
    types = criteria_types_for(criteria)
    weights = {c: 1.0 / len(criteria) for c in criteria}
    scale_vec = np.asarray([3.0, 0.5, 7.0, 1.75, 4.0], dtype=float)[: len(criteria)]

    for method in me.CLASSICAL_MCDM_METHODS:
        for df in (positive_df, mixed_df):
            table, details = me.rank_alternatives(df, criteria, types, weights, method)
            scores = pd.to_numeric(table["Skor"], errors="coerce").to_numpy(dtype=float)
            require(np.isfinite(scores).all(), f"{method}: non-finite scores.")
            require(table["Alternatif"].nunique() == len(df), f"{method}: duplicate alternatives in rank table.")
            require(isinstance(details, dict), f"{method}: details payload missing.")

        scaled = mixed_df.mul(scale_vec, axis=1)
        base_table, _ = me.rank_alternatives(mixed_df, criteria, types, weights, method)
        scaled_table, _ = me.rank_alternatives(scaled, criteria, types, weights, method)
        require(
            list(base_table["Alternatif"]) == list(scaled_table["Alternatif"]),
            f"{method}: ranking changed under positive rescaling.",
        )

    for method in me.FUZZY_MCDM_METHODS:
        crisp_method = method.replace("Fuzzy ", "")
        fuzzy_table, _ = me.rank_alternatives(positive_df, criteria, types, weights, method, fuzzy_spread=0.0)
        crisp_table, _ = me.rank_alternatives(positive_df, criteria, types, weights, crisp_method, fuzzy_spread=0.0)
        fuzzy_scores = pd.to_numeric(fuzzy_table["Skor"], errors="coerce").to_numpy(dtype=float)
        crisp_scores = pd.to_numeric(crisp_table["Skor"], errors="coerce").to_numpy(dtype=float)
        require(
            list(fuzzy_table["Alternatif"]) == list(crisp_table["Alternatif"]),
            f"{method}: spread=0 did not preserve the crisp ranking.",
        )
        require(np.allclose(fuzzy_scores, crisp_scores, atol=1e-10), f"{method}: spread=0 scores diverged from crisp scores.")


def test_analysis_bundles() -> None:
    df = make_dataset(505, mixed=False, n_alt=10, n_crit=5)
    criteria = list(df.columns)
    types = criteria_types_for(criteria)

    cfg = me.AnalysisConfig(
        criteria=criteria,
        criteria_types=types,
        weight_method="CRITIC",
        ranking_method="VIKOR",
        compare_methods=["TOPSIS", "PROMETHEE"],
        sensitivity_iterations=180,
        sensitivity_sigma=0.10,
        run_heavy_robustness=True,
    )
    result = me.run_full_analysis(df, cfg)
    require(result["weights"]["table"].shape[0] == len(criteria), "Objective bundle: weight table size mismatch.")
    require(isinstance(result["ranking"]["details"], dict), "Objective bundle: ranking details missing.")
    require(bool(result.get("decision_confidence")), "Objective bundle: decision confidence missing.")
    require(isinstance(result["comparison"].get("spearman_matrix"), pd.DataFrame), "Objective bundle: comparison matrix missing.")
    require(isinstance(result["sensitivity"].get("monte_carlo_summary"), pd.DataFrame), "Objective bundle: sensitivity summary missing.")
    require("Bulgular" in result["report_sections"], "Objective bundle: report sections missing findings.")

    cfg_manual = me.AnalysisConfig(
        criteria=criteria,
        criteria_types=types,
        weight_method="Entropy",
        weight_mode="manual",
        manual_weights={c: idx + 1.0 for idx, c in enumerate(criteria)},
        ranking_method="TOPSIS",
        compare_methods=["VIKOR"],
        sensitivity_iterations=0,
        run_heavy_robustness=False,
    )
    manual_result = me.run_full_analysis(df, cfg_manual)
    require(manual_result["weights"]["details"]["mode"] == "manual", "Manual bundle: mode not preserved.")
    require(manual_result["ranking"]["table"] is not None, "Manual bundle: ranking table missing.")
    require(manual_result["decision_confidence"]["verdict"] in {"high", "medium", "low"}, "Manual bundle: invalid confidence verdict.")


def run_single_sample_flow() -> None:
    at = make_app_test(120)
    at.run()
    assert_no_streamlit_exceptions(at, "single:init")
    click_button_by_key(at, "btn_sample_tr_main")
    click_button_by_key(at, "btn_use_upload_data_main")
    click_button_by_key(at, "step1_continue_btn")
    click_button_by_label(at, "Veri Ön İşleme Bitti", "Preprocessing Complete")
    set_checkbox(at, "weight_cb_CRITIC", True)
    set_checkbox(at, "rank_cb_VIKOR", True)
    set_checkbox(at, "rank_cb_TOPSIS", True)
    set_checkbox(at, "run_heavy_robustness", True)
    click_button_by_label(at, "Analiz Zamanı", "Run Analysis")
    result = at.session_state["analysis_result"]
    require(isinstance(result, dict), "Single flow: analysis result missing.")
    _, sheetnames = get_download_workbook(at)
    required = {"Kapak", "Ham Veri", "CRITIC Analizi", "Özet Bulgular"}
    require(required.issubset(set(sheetnames)), f"Single flow: required base sheets missing. Found {sheetnames}")
    require(any(name.endswith("Sonuçları") for name in sheetnames), f"Single flow: no ranking results sheet found. Found {sheetnames}")
    require("Yöntem Karşılaştırması" in sheetnames, f"Single flow: comparison sheet missing. Found {sheetnames}")
    require("Duyarlılık Analizi" in sheetnames, f"Single flow: sensitivity sheet missing. Found {sheetnames}")
    method_tables = result.get("comparison", {}).get("method_tables") or {}
    require("VIKOR" in method_tables, "Single flow: VIKOR comparison output missing.")


def run_manual_flow() -> None:
    at = make_app_test(120)
    at.run()
    assert_no_streamlit_exceptions(at, "manual:init")
    at.radio(key="data_entry_mode").set_value("✍️ Manuel Giriş").run()
    assert_no_streamlit_exceptions(at, "manual:mode")
    click_button_by_key(at, "btn_fill_manual_sample_main")
    click_button_by_key(at, "btn_use_manual_data_from_paste_main")
    require(at.session_state["data_source_id"] == "manual_entry", "Manual flow: data source did not switch to manual.")
    click_button_by_key(at, "step1_continue_btn")
    click_button_by_label(at, "Veri Ön İşleme Bitti", "Preprocessing Complete")
    set_checkbox(at, "weight_cb_CRITIC", True)
    set_checkbox(at, "rank_cb_TOPSIS", True)
    click_button_by_label(at, "Analiz Zamanı", "Run Analysis")
    result = at.session_state["analysis_result"]
    require(isinstance(result, dict), "Manual flow: analysis result missing.")
    _, sheetnames = get_download_workbook(at)
    require("Ham Veri" in sheetnames or "Raw Data" in sheetnames, f"Manual flow: raw data sheet missing: {sheetnames}")


def run_panel_flow() -> None:
    at = make_app_test(150)
    at.run()
    assert_no_streamlit_exceptions(at, "panel:init")
    click_button_by_key(at, "btn_sample_panel_en_main")
    click_button_by_key(at, "btn_use_upload_data_main")
    click_button_by_key(at, "panel_years_select_all")
    click_button_by_key(at, "step1_continue_btn")
    click_button_by_label(at, "Preprocessing Complete", "Veri Ön İşleme Bitti")
    set_checkbox(at, "weight_cb_CRITIC", True)
    set_checkbox(at, "rank_cb_TOPSIS", True)
    set_checkbox(at, "rank_cb_VIKOR", True)
    click_button_by_label(at, "Run Analysis", "Analiz Zamanı")
    panel_results = at.session_state["panel_results"]
    require(isinstance(panel_results, dict), "Panel flow: panel results missing.")
    require(len(panel_results) >= 2, "Panel flow: expected multiple yearly results.")
    _, sheetnames = get_download_workbook(at)
    tr_required = {"Kapak", "Ham Veri", "Panel_Ozet", "Panel_Matris", "Özet Bulgular"}
    en_required = {"Cover", "Raw Data", "Panel_Summary", "Panel_Matrix", "Summary Findings"}
    require(
        tr_required.issubset(set(sheetnames)) or en_required.issubset(set(sheetnames)),
        f"Panel flow: expected sheets missing. Found {sheetnames}",
    )


def run_fuzzy_single_flow() -> None:
    at = make_app_test(150)
    at.run()
    assert_no_streamlit_exceptions(at, "fuzzy_single:init")
    at.radio(key="data_value_mode").set_value("🔺 Üçgensel bulanık (TFN)").run()
    assert_no_streamlit_exceptions(at, "fuzzy_single:value_mode")
    click_button_by_key(at, "btn_sample_en_main")
    click_button_by_key(at, "btn_use_upload_data_main")
    require(
        (session_state_value(at, "input_data_profile", {}) or {}).get("value_mode") == "tfn",
        "Fuzzy single flow: TFN profile not preserved.",
    )
    click_button_by_key(at, "step1_continue_btn")
    click_button_by_label(at, "Veri Ön İşleme Bitti", "Preprocessing Complete")
    set_checkbox(at, "weight_cb_Fuzzy Entropy", True)
    set_radio_by_label(at, ["Analiz Katmanı", "Analysis Layer"], "İleri Düzey Yöntemler (Fuzzy)")
    set_checkbox(at, "rank_cb_Fuzzy TOPSIS", True)
    click_button_by_label(at, "Analiz Zamanı", "Run Analysis")
    result = at.session_state["analysis_result"]
    require(isinstance(result, dict), "Fuzzy single flow: analysis result missing.")
    require(result["weights"]["method"] == "Fuzzy Entropy", "Fuzzy single flow: fuzzy weight method not preserved.")
    require(result["ranking"]["method"] == "Fuzzy TOPSIS", "Fuzzy single flow: fuzzy ranking method not preserved.")
    require((session_state_value(at, "input_data_profile", {}) or {}).get("value_mode") == "tfn", "Fuzzy single flow: input profile lost after run.")


def run_fuzzy_panel_flow() -> None:
    at = make_app_test(180)
    at.run()
    assert_no_streamlit_exceptions(at, "fuzzy_panel:init")
    at.radio(key="data_value_mode").set_value("🔺 Üçgensel bulanık (TFN)").run()
    assert_no_streamlit_exceptions(at, "fuzzy_panel:value_mode")
    click_button_by_key(at, "btn_sample_panel_en_main")
    click_button_by_key(at, "btn_use_upload_data_main")
    require(
        (session_state_value(at, "input_data_profile", {}) or {}).get("value_mode") == "tfn",
        "Fuzzy panel flow: TFN profile not preserved.",
    )
    click_button_by_key(at, "panel_years_select_all")
    click_button_by_key(at, "step1_continue_btn")
    click_button_by_label(at, "Preprocessing Complete", "Veri Ön İşleme Bitti")
    set_checkbox(at, "weight_cb_Fuzzy CRITIC", True)
    set_radio_by_label(at, ["Analiz Katmanı", "Analysis Layer"], "İleri Düzey Yöntemler (Fuzzy)")
    set_checkbox(at, "rank_cb_Fuzzy TOPSIS", True)
    click_button_by_label(at, "Run Analysis", "Analiz Zamanı")
    panel_results = at.session_state["panel_results"]
    require(isinstance(panel_results, dict) and panel_results, "Fuzzy panel flow: panel results missing.")
    first_result = next(iter(panel_results.values()))
    require(first_result["weights"]["method"] == "Fuzzy CRITIC", "Fuzzy panel flow: fuzzy weight method not preserved.")
    require(first_result["ranking"]["method"] == "Fuzzy TOPSIS", "Fuzzy panel flow: fuzzy ranking method not preserved.")


def run_test(name: str, fn: Callable[[], None]) -> tuple[str, bool, str]:
    try:
        fn()
        return name, True, ""
    except Exception as exc:
        tb = traceback.format_exc()
        return name, False, f"{exc}\n{tb}"


def main() -> int:
    tests: list[tuple[str, Callable[[], None]]] = [
        ("weight_methods", test_weight_methods),
        ("ranking_methods", test_ranking_methods),
        ("analysis_bundles", test_analysis_bundles),
        ("app_single_flow", run_single_sample_flow),
        ("app_manual_flow", run_manual_flow),
        ("app_panel_flow", run_panel_flow),
        ("app_fuzzy_single_flow", run_fuzzy_single_flow),
        ("app_fuzzy_panel_flow", run_fuzzy_panel_flow),
    ]

    failures: list[tuple[str, str]] = []
    for name, fn in tests:
        test_name, ok, details = run_test(name, fn)
        status = "PASS" if ok else "FAIL"
        print(f"[{status}] {test_name}")
        if not ok:
            failures.append((test_name, details))
            print(details)

    print(f"\nSummary: {len(tests) - len(failures)}/{len(tests)} checks passed.")
    if failures:
        print("Failed checks:")
        for name, _ in failures:
            print(f"- {name}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
